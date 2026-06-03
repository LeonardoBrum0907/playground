#!/usr/bin/env python3
"""
Automatiza a transformacao de apontamentos em um arquivo Excel padrao.

Entrada:
  - CSV com as colunas originais exportadas.
Saida:
  - XLSX com duas abas:
      1) Dados  -> apenas 5 colunas selecionadas
      2) Resumo -> total de horas por colaborador (formula automatica)
"""

from __future__ import annotations

import argparse
import csv
import sys
import unicodedata
from datetime import date, datetime, time, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ModuleNotFoundError:
    print(
        "Dependencia ausente: openpyxl.\n"
        "Instale com: python3 -m pip install openpyxl",
        file=sys.stderr,
    )
    raise

HEADER_COLOR = "1F4E78"
TARGET_HEADERS = [
    "Atividade",
    "Horas Apontadas",
    "Apontado por",
    "Data de Início Real",
    "Data de Fim Real",
]

SOURCE_ALIASES = {
    "Atividade": ["atividade"],
    "Horas Apontadas": ["horas apontadas", "horas apontado"],
    "Apontado por": ["apontado por"],
    "Data de Início Real": ["data de inicio real"],
    "Data de Fim Real": ["data de fim real"],
}


def normalize(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    without_accents = "".join(c for c in normalized if not unicodedata.combining(c))
    return without_accents.strip().lower()


def parse_hms_to_excel_time(value: str) -> float:
    if not value or value.strip() in {"", "-"}:
        return 0.0

    parts = value.strip().split(":")
    if len(parts) != 3:
        return 0.0

    try:
        hours, minutes, seconds = (int(p) for p in parts)
    except ValueError:
        return 0.0

    total_seconds = hours * 3600 + minutes * 60 + seconds
    return total_seconds / 86400.0


def parse_excel_time_to_fraction(value: object) -> float:
    if value is None:
        return 0.0

    if isinstance(value, bool):
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, timedelta):
        return value.total_seconds() / 86400.0

    if isinstance(value, time):
        total_seconds = value.hour * 3600 + value.minute * 60 + value.second
        return total_seconds / 86400.0

    if isinstance(value, str):
        return parse_hms_to_excel_time(value)

    return 0.0


def normalize_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y, %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def detect_dialect(csv_path: Path) -> csv.Dialect:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        sample = file.read(4096)
        file.seek(0)
        try:
            return csv.Sniffer().sniff(sample, delimiters="\t,;")
        except csv.Error:
            return csv.excel_tab


def resolve_source_headers(fieldnames: list[str]) -> dict[str, str]:
    normalized_headers = {normalize(name): name for name in fieldnames}
    selected: dict[str, str] = {}

    for target_header in TARGET_HEADERS:
        aliases = SOURCE_ALIASES[target_header]
        source_header = None
        for alias in aliases:
            source_header = normalized_headers.get(alias)
            if source_header:
                break
        if source_header is None:
            aliases_text = ", ".join(aliases)
            raise ValueError(
                f"Coluna obrigatoria nao encontrada para '{target_header}'. "
                f"Aliases esperados: {aliases_text}"
            )
        selected[target_header] = source_header

    return selected


def apply_header_style(worksheet, total_columns: int) -> None:
    for column_index in range(1, total_columns + 1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color=HEADER_COLOR,
            end_color=HEADER_COLOR,
            fill_type="solid",
        )


def parse_rows_from_input_file(input_path: Path) -> list[dict[str, str | float]]:
    dialect = detect_dialect(input_path)
    rows: list[dict[str, str | float]] = []

    with input_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError(f"CSV sem cabecalho: {input_path}")

        source_headers = resolve_source_headers(reader.fieldnames)

        for row in reader:
            entry = {
                "Atividade": (row.get(source_headers["Atividade"]) or "").strip(),
                "Horas Apontadas": parse_hms_to_excel_time(
                    row.get(source_headers["Horas Apontadas"]) or ""
                ),
                "Apontado por": (row.get(source_headers["Apontado por"]) or "").strip(),
                "Data de Início Real": (row.get(source_headers["Data de Início Real"]) or "").strip(),
                "Data de Fim Real": (row.get(source_headers["Data de Fim Real"]) or "").strip(),
            }
            rows.append(entry)

    return rows


def write_output_workbook(rows: list[dict[str, str | float]], output_path: Path) -> None:
    workbook = Workbook()
    ws_data = workbook.active
    ws_data.title = "Dados"
    ws_data.append(TARGET_HEADERS)

    for index, row in enumerate(rows, start=2):
        ws_data.cell(row=index, column=1, value=row["Atividade"])
        hours_cell = ws_data.cell(row=index, column=2, value=row["Horas Apontadas"])
        hours_cell.number_format = "[h]:mm:ss"
        ws_data.cell(row=index, column=3, value=row["Apontado por"])
        ws_data.cell(row=index, column=4, value=row["Data de Início Real"])
        ws_data.cell(row=index, column=5, value=row["Data de Fim Real"])

    apply_header_style(ws_data, total_columns=5)
    ws_data.column_dimensions["A"].width = 55
    ws_data.column_dimensions["B"].width = 16
    ws_data.column_dimensions["C"].width = 16
    ws_data.column_dimensions["D"].width = 20
    ws_data.column_dimensions["E"].width = 20

    ws_summary = workbook.create_sheet("Resumo")
    summary_headers = ["Colaborador", "Total de Horas Apontadas"]
    ws_summary.append(summary_headers)

    collaborators = sorted(
        {row["Apontado por"] for row in rows if row["Apontado por"]},
        key=lambda name: str(name).lower(),
    )

    for row_index, collaborator in enumerate(collaborators, start=2):
        ws_summary.cell(row=row_index, column=1, value=collaborator)
        formula = f"=SUMIF(Dados!$C:$C,A{row_index},Dados!$B:$B)"
        total_cell = ws_summary.cell(row=row_index, column=2, value=formula)
        total_cell.number_format = "[h]:mm:ss"

    apply_header_style(ws_summary, total_columns=2)
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 24

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def row_signature(row: dict[str, str | float]) -> tuple[str, float, str, str, str]:
    return (
        str(row["Atividade"]).strip(),
        round(float(row["Horas Apontadas"]), 10),
        str(row["Apontado por"]).strip(),
        str(row["Data de Início Real"]).strip(),
        str(row["Data de Fim Real"]).strip(),
    )


def load_existing_rows(output_path: Path) -> list[dict[str, str | float]]:
    if not output_path.exists():
        return []

    workbook = load_workbook(output_path, data_only=False)
    if "Dados" not in workbook.sheetnames:
        return []

    ws_data = workbook["Dados"]
    if ws_data.max_row < 2:
        return []

    headers = [ws_data.cell(row=1, column=col).value for col in range(1, 6)]
    normalized_headers = {
        normalize(str(header)): idx
        for idx, header in enumerate(headers, start=1)
        if header is not None
    }

    required = {
        "atividade": "Atividade",
        "horas apontadas": "Horas Apontadas",
        "apontado por": "Apontado por",
        "data de inicio real": "Data de Início Real",
        "data de fim real": "Data de Fim Real",
    }
    missing = [name for name in required if name not in normalized_headers]
    if missing:
        return []

    rows: list[dict[str, str | float]] = []
    for row_index in range(2, ws_data.max_row + 1):
        atividade = normalize_cell_text(ws_data.cell(row=row_index, column=normalized_headers["atividade"]).value)
        horas = parse_excel_time_to_fraction(
            ws_data.cell(row=row_index, column=normalized_headers["horas apontadas"]).value
        )
        apontado_por = normalize_cell_text(
            ws_data.cell(row=row_index, column=normalized_headers["apontado por"]).value
        )
        data_inicio = normalize_cell_text(
            ws_data.cell(row=row_index, column=normalized_headers["data de inicio real"]).value
        )
        data_fim = normalize_cell_text(
            ws_data.cell(row=row_index, column=normalized_headers["data de fim real"]).value
        )

        if not any([atividade, horas, apontado_por, data_inicio, data_fim]):
            continue

        rows.append(
            {
                "Atividade": atividade,
                "Horas Apontadas": horas,
                "Apontado por": apontado_por,
                "Data de Início Real": data_inicio,
                "Data de Fim Real": data_fim,
            }
        )

    return rows


def select_input_files(input_file: Path | None, input_dir: Path, process_all: bool) -> list[Path]:
    if input_file:
        if not input_file.exists():
            raise FileNotFoundError(f"Arquivo nao encontrado: {input_file}")
        return [input_file]

    input_dir.mkdir(parents=True, exist_ok=True)
    input_files = [
        *input_dir.glob("*.csv"),
        *input_dir.glob("*.txt"),
    ]
    input_files_sorted = sorted(input_files, key=lambda path: path.stat().st_mtime, reverse=True)
    if not input_files_sorted:
        raise FileNotFoundError(
            f"Nenhum CSV/TXT encontrado em '{input_dir}'. "
            "Coloque o arquivo na pasta e execute novamente."
        )

    if process_all:
        return list(reversed(input_files_sorted))
    return [input_files_sorted[0]]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Transforma CSV de apontamentos em XLSX com aba de resumo automatico."
    )
    parser.add_argument(
        "--arquivo",
        type=Path,
        default=None,
        help="Caminho de um CSV especifico para processar.",
    )
    parser.add_argument(
        "--entrada",
        type=Path,
        default=Path("entrada"),
        help="Pasta com os CSVs (padrao: ./entrada).",
    )
    parser.add_argument(
        "--saida",
        type=Path,
        default=Path("saida"),
        help="Pasta de saida para os XLSX (padrao: ./saida).",
    )
    parser.add_argument(
        "--todos",
        action="store_true",
        help="Processa todos os CSVs/TXTs da pasta de entrada.",
    )
    parser.add_argument(
        "--acumular-em",
        type=Path,
        default=None,
        help=(
            "Acumula dados em um unico XLSX historico, sem apagar os dias anteriores. "
            "Nao duplica linhas identicas."
        ),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        files = select_input_files(args.arquivo, args.entrada, args.todos)
        if args.acumular_em:
            historical_rows = load_existing_rows(args.acumular_em)
            existing_signatures = {row_signature(row) for row in historical_rows}
            new_rows = 0

            for input_file in files:
                parsed_rows = parse_rows_from_input_file(input_file)
                for row in parsed_rows:
                    signature = row_signature(row)
                    if signature in existing_signatures:
                        continue
                    historical_rows.append(row)
                    existing_signatures.add(signature)
                    new_rows += 1

            write_output_workbook(historical_rows, args.acumular_em)
            print(
                f"OK: {len(historical_rows)} registros totais ({new_rows} novos) "
                f"-> {args.acumular_em}"
            )
            return 0

        for input_file in files:
            output_name = f"{input_file.stem}_editada.xlsx"
            output_file = args.saida / output_name
            rows = parse_rows_from_input_file(input_file)
            write_output_workbook(rows, output_file)
            print(f"OK: {input_file} -> {output_file}")
        return 0
    except Exception as exc:  # noqa: BLE001
        print(f"Erro: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
